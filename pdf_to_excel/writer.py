"""Excel file writer with Thai language support."""

from pathlib import Path
from typing import List, Union

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .exceptions import WriterError


class ExcelWriter:
    """Write data to Excel files with proper formatting."""

    def __init__(self, thai_font: str = "TH SarabunPSK", thai_font_size: int = 16):
        """
        Initialize Excel writer.

        Args:
            thai_font: Font name for Thai text (default: TH SarabunPSK)
            thai_font_size: Font size for Thai text (default: 16)
        """
        self.thai_font = thai_font
        self.thai_font_size = thai_font_size

    def write_table_to_excel(
        self,
        table_data: List[List[str]],
        output_path: Union[str, Path],
        apply_formatting: bool = True,
    ) -> None:
        """
        Write table data to Excel file.

        Args:
            table_data: List of rows (each row is a list of cells)
            output_path: Path to output Excel file
            apply_formatting: Whether to apply formatting (headers, fonts, etc.)

        Raises:
            WriterError: If writing fails
        """
        try:
            # Create output directory if needed
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)

            # Find the row with most columns (likely the header)
            max_cols = max(len(row) for row in table_data)

            # Pad rows to have consistent column count
            padded_data = []
            for row in table_data:
                if len(row) < max_cols:
                    row = row + [""] * (max_cols - len(row))
                padded_data.append(row)

            # Convert to DataFrame
            if len(padded_data) > 1:
                # First row with max columns as headers
                df = pd.DataFrame(padded_data[1:], columns=padded_data[0])
            else:
                # No headers
                df = pd.DataFrame(padded_data)

            # Write to Excel
            df.to_excel(str(output_path), index=False, engine="openpyxl")

            # Apply formatting if requested
            if apply_formatting:
                self._apply_formatting(output_path)

        except Exception as e:
            raise WriterError(f"Failed to write Excel file: {e}")

    def _apply_formatting(self, excel_path: Path) -> None:
        """
        Apply formatting to Excel file (Thai font, bold headers, etc.).

        Args:
            excel_path: Path to Excel file
        """
        try:
            # Load the workbook
            wb = load_workbook(excel_path)
            ws = wb.active

            # Header formatting (first row)
            header_font = Font(
                name=self.thai_font, size=self.thai_font_size, bold=True
            )
            header_fill = PatternFill(
                start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
            )

            # Apply to header row
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Data formatting (Thai font for all cells)
            data_font = Font(name=self.thai_font, size=self.thai_font_size)

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = data_font
                    cell.alignment = Alignment(vertical="center")

            # Auto-adjust column widths (accounting for Thai characters)
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if cell.value:
                            # Thai characters need more width
                            cell_value = str(cell.value)
                            # Count Thai chars (roughly 1.5x wider than ASCII)
                            thai_chars = sum(
                                1 for c in cell_value if "\u0E00" <= c <= "\u0E7F"
                            )
                            ascii_chars = len(cell_value) - thai_chars
                            effective_length = thai_chars * 1.5 + ascii_chars

                            if effective_length > max_length:
                                max_length = effective_length
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save
            wb.save(excel_path)

        except Exception as e:
            # Non-critical error, file is still usable
            print(f"Warning: Could not apply formatting: {e}")
