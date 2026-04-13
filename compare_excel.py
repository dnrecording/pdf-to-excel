#!/usr/bin/env python3
"""Compare two Excel files to see formatting and content differences."""

import pandas as pd
from openpyxl import load_workbook
import sys

def compare_excel_files(file1, file2):
    """Compare two Excel files."""
    print("=" * 80)
    print("Comparing Excel Files")
    print("=" * 80)
    print(f"File 1 (Windows): {file1}")
    print(f"File 2 (macOS):   {file2}")
    print()

    # Read with pandas to compare data
    print("-" * 80)
    print("DATA COMPARISON (using pandas)")
    print("-" * 80)

    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)

    print(f"\nWindows DataFrame shape: {df1.shape}")
    print(f"macOS DataFrame shape:   {df2.shape}")

    print("\nWindows DataFrame head:")
    print(df1.head())

    print("\nmacOS DataFrame head:")
    print(df2.head())

    # Check if data is identical
    if df1.equals(df2):
        print("\n✅ Data content is IDENTICAL")
    else:
        print("\n❌ Data content is DIFFERENT")

        # Show differences
        print("\nDifferences:")
        if df1.shape != df2.shape:
            print(f"  - Different shapes: {df1.shape} vs {df2.shape}")

        if list(df1.columns) != list(df2.columns):
            print(f"  - Different columns:")
            print(f"    Windows: {list(df1.columns)}")
            print(f"    macOS:   {list(df2.columns)}")

        # Compare cell by cell for first few rows
        print("\nCell-by-cell comparison (first 5 rows):")
        for i in range(min(5, len(df1), len(df2))):
            for col in df1.columns:
                if col in df2.columns:
                    val1 = df1.iloc[i][col]
                    val2 = df2.iloc[i][col]
                    if val1 != val2:
                        print(f"  Row {i}, Column '{col}':")
                        print(f"    Windows: {repr(val1)}")
                        print(f"    macOS:   {repr(val2)}")

    # Read with openpyxl to compare formatting
    print("\n" + "-" * 80)
    print("FORMATTING COMPARISON (using openpyxl)")
    print("-" * 80)

    wb1 = load_workbook(file1)
    wb2 = load_workbook(file2)

    ws1 = wb1.active
    ws2 = wb2.active

    print(f"\nWindows worksheet dimensions: {ws1.dimensions}")
    print(f"macOS worksheet dimensions:   {ws2.dimensions}")

    # Compare fonts, column widths, etc.
    print("\nColumn widths:")
    for col in range(1, min(10, ws1.max_column + 1)):
        col_letter = chr(64 + col)
        width1 = ws1.column_dimensions[col_letter].width
        width2 = ws2.column_dimensions[col_letter].width
        if width1 != width2:
            print(f"  Column {col_letter}: Windows={width1}, macOS={width2}")

    # Check first few cells for font differences
    print("\nFont comparison (first row):")
    for col in range(1, min(5, ws1.max_column + 1)):
        cell1 = ws1.cell(1, col)
        cell2 = ws2.cell(1, col)

        if cell1.font.name != cell2.font.name:
            print(f"  Column {col}: Windows font='{cell1.font.name}', macOS font='{cell2.font.name}'")
        if cell1.font.size != cell2.font.size:
            print(f"  Column {col}: Windows size={cell1.font.size}, macOS size={cell2.font.size}")
        if cell1.font.bold != cell2.font.bold:
            print(f"  Column {col}: Windows bold={cell1.font.bold}, macOS bold={cell2.font.bold}")

    print("\n" + "=" * 80)

if __name__ == "__main__":
    file1 = "/Users/wsrisuntikan/Downloads/test_windows.xlsx"
    file2 = "/Users/wsrisuntikan/Downloads/test_mac.xlsx"

    compare_excel_files(file1, file2)
